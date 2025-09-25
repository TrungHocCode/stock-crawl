import openpyxl
import json
from datetime import datetime
from openpyxl.styles import PatternFill
import os

class ExcelProcessor:
    def __init__ (self, json_file="./data/processed"):
        self.json_file=json_file
        self.all_rows = []
    def load_json(self):
        try:
            with open(self.json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Lỗi khi đọc file JSON {self.json_file}: {e}")
            return {}
        return data

    def mark_price_change(self):
        for i in range(1, len(self.all_rows)):
            change = False
            direction = None
            has_missing_data = False

            # Kiểm tra dữ liệu đủ 3 mức giá mua/bán
            for side in ['mua', 'bán']:
                for pos in range(1, 4):
                    if not self.all_rows[i].get(f"Giá chờ {side} {pos}", ""):
                        has_missing_data = True
                        break
                if has_missing_data:
                    break

            if not has_missing_data:
                for side in ['mua', 'bán']:
                    for pos in range(1, 4):
                        prev_val = self.all_rows[i - 1].get(f"Giá chờ {side} {pos}", "")
                        curr_val = self.all_rows[i].get(f"Giá chờ {side} {pos}", "")
                        if prev_val != curr_val and prev_val and curr_val:
                            change = True
                            try:
                                direction = "decrease" if float(curr_val) < float(prev_val) else "increase"
                            except:
                                direction = "unknown"
                            break
                    if change:
                        break

            if change:
                self.all_rows[i]["Thay đổi bước giá"] = 1
                self.all_rows[i]["price_direction"] = direction

    def calculate_diff_for_side(self, i, side):
        if i == 0:
            return [0] * 3

        prev_row = self.all_rows[i - 1]
        current_row = self.all_rows[i]

        if current_row.get("Thay đổi bước giá") == 1:
            prev_prices = [float(prev_row.get(f"Giá chờ {side} {j+1}", 0)) for j in range(3)]
            diffs = []
            for j in range(3):
                curr_price = float(current_row.get(f"Giá chờ {side} {j+1}", 0))
                curr_vol = int(current_row.get(f"Chờ {side} {j+1}", 0))

                matched_idx = -1
                for k, prev_price in enumerate(prev_prices):
                    if abs(curr_price - prev_price) <= 0.001:
                        matched_idx = k
                        break

                if matched_idx != -1:
                    prev_vol_same_index = int(prev_row.get(f"Chờ {side} {matched_idx+1}", 0))
                    diff = curr_vol - prev_vol_same_index
                else:
                    diff = curr_vol

                diffs.append(diff)
            return diffs
        return [
            int(current_row.get(f"Chờ {side} {j+1}", 0)) - int(prev_row.get(f"Chờ {side} {j+1}", 0))
            for j in range(3)
        ]

    def detect_ban_quet_by_delta(self, i, window=10, delta_threshold=0.05):
        if i < window:
            return False

        deltas = []
        for j in range(i - window, i):
            row = self.all_rows[j]
            gia_mua_1 = float(row.get("Giá chờ mua 1", 0))
            gia_ban_1 = float(row.get("Giá chờ bán 1", 0))
            if gia_mua_1 == 0 or gia_ban_1 == 0:
                continue
            deltas.append(gia_ban_1 - gia_mua_1)

        if len(deltas) < 3:
            return False

        current_row = self.all_rows[i]
        current_delta = float(current_row.get("Giá chờ bán 1", 0)) - float(current_row.get("Giá chờ mua 1", 0))
        avg_prev_delta = sum(deltas) / len(deltas)

        if float(current_row.get("Giá chờ mua 1")) - float(self.all_rows[i - 1].get("Giá chờ mua 1")) == 0:
            return False

        return current_delta - avg_prev_delta > delta_threshold
    
    def is_zone_empty(self,zone):
        """Zone rỗng nếu mọi giá trị đều là chuỗi rỗng"""
        if not zone:
            return True
        for z in zone:
            if any(v != "" for v in z.values()):
                return False
        return True
    
    def preprocess_rows(self, data):
        timestamps = sorted(data.keys())
        last_zone = None   # lưu lại zone gần nhất (nếu không rỗng)
        last_row = None

        for timestamp in timestamps:
            entry = data[timestamp]
            zone = entry.get('zone', [])
            # Parse timestamp
            try:
                dt_obj = datetime.strptime(timestamp, "%H:%M:%S.%f")
            except ValueError:
                dt_obj = datetime.strptime(timestamp, "%H:%M:%S")
            timestamp_formatted = dt_obj.strftime("%H:%M:%S")

            row = {
                'Thời gian chi tiết': timestamp,
                'Thời gian': timestamp_formatted,
                'Giá': entry.get('Giá', 0),
                'Khối lượng': entry.get('KL', 0),
                'Loại': entry.get('M/B', 0),
                'Khớp lần 1': entry.get('Khớp', 0)
            }
            if self.is_zone_empty(zone):
                zone=last_zone
            else:
                last_zone=zone
            for i in range(3):
                zone_data = zone[i] if i < len(zone) else {}
                row[f'Chờ mua {i+1}'] = int(zone_data.get(f'KL_mua {i+1}', 0))
                row[f'Giá chờ mua {i+1}'] = zone_data.get(f'Gia_mua {i+1}', 0)
                row[f'Chờ bán {i+1}'] = int(zone_data.get(f'KL_ban {i+1}', 0))
                row[f'Giá chờ bán {i+1}'] = zone_data.get(f'Gia_ban {i+1}', 0)

            row_without_time = {k: v for k, v in row.items() if k != "Thời gian"}

            if last_row is not None and row_without_time == last_row:
                continue  # bỏ qua vì trùng dữ liệu

            self.all_rows.append(row)
            last_row = row_without_time
        
    def process(self):
        self.mark_price_change()
        for i in range(len(self.all_rows)):
            for side in ['mua', 'bán']:
                diffs = self.calculate_diff_for_side(i, side)
                for j in range(3):
                    self.all_rows[i][f"Tăng/giảm chờ {side} {j+1}"] = diffs[j]

            if self.detect_ban_quet_by_delta(i):
                self.all_rows[i]["Bán quét"] = "1"
            else:
                self.all_rows[i]["Bán quét"] = ""

    def export_to_excel(self):
        stock_name = os.path.basename(self.json_file).split('.')[0]
        excel_file = os.path.join(os.path.dirname(self.json_file), f"{stock_name}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dữ liệu"

        headers = [
            "Thời gian", "Giá", "Khối lượng", "Loại",
            "Tăng/giảm chờ mua 1", "Tăng/giảm chờ bán 1",
            "Tăng/giảm chờ mua 2", "Tăng/giảm chờ bán 2",
            "Tăng/giảm chờ mua 3", "Tăng/giảm chờ bán 3",
            "Chờ mua 1", "Chờ bán 1", "Chờ mua 2", "Chờ bán 2",
            "Chờ mua 3", "Chờ bán 3",
            "Giá chờ mua 1", "Giá chờ bán 1",
            "Giá chờ mua 2", "Giá chờ bán 2",
            "Giá chờ mua 3", "Giá chờ bán 3",
            "Thay đổi bước giá", "Bán quét"
        ]
        ws.append(headers)

        for row in sorted(self.all_rows, key=lambda x: x['Thời gian chi tiết'], reverse=True):
            ws.append([row.get(header, '') for header in headers])

        color_increase = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        color_decrease = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        dark_red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
        for r_idx, row in enumerate(sorted(self.all_rows, key=lambda x: x['Thời gian chi tiết'], reverse=True), start=2):
            if str(row.get("Thay đổi bước giá")) == "1":
                direction = row.get("price_direction", "decrease")
                fill_color = color_decrease if direction == "decrease" else color_increase
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_color

            if str(row.get("Bán quét")) == "1":
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = dark_red_fill
        wb.save(excel_file)
        print(f"Đã xuất ra Excel: {excel_file}")

    def run(self):
        data = self.load_json()
        if not data:
            return
        self.preprocess_rows(data)
        self.process()
        self.export_to_excel()